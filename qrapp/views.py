from django.shortcuts import render, redirect
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from .models import FailedLoginAttempts, Register
from django.contrib.auth import authenticate, login
import uuid
import json
from django.http import JsonResponse,HttpResponse
from django.views.decorators.csrf import csrf_exempt
import pytz
from datetime import datetime
from django.utils import timezone
from openpyxl import Workbook
from io import BytesIO 
from phonenumbers import format_number, PhoneNumberFormat, parse, NumberParseException

MAX_FAILED_ATTEMPTS = 3

def set_device_cookie(response, device_id):
    response.set_cookie('device_id', device_id, max_age=365 * 24 * 60 * 60)

def get_device_id(request):
    return request.COOKIES.get('device_id', str(uuid.uuid4()))

def is_device_blocked(device_id):
    try:
        failed_attempt = FailedLoginAttempts.objects.get(device_id=device_id)
        if failed_attempt.is_active:
            return True
    except FailedLoginAttempts.DoesNotExist:
        return False
    return False

def adminlogin(request):
    if request.method == 'POST':
        device_id = get_device_id(request)
        if is_device_blocked(device_id):
            return render(request, 'adminlogin.html', {'blocked': True, 
                                                        'error_message': 'Your device is permanently blocked due to multiple failed login attempts.'})

        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.is_superuser or (user.is_staff and not user.is_active):
                FailedLoginAttempts.objects.filter(device_id=device_id).update(attempts=0, is_active=False)
                response = redirect('adminpage')
                set_device_cookie(response, device_id)
                login(request, user)
                return response
            else:
                return render(request, 'adminlogin.html', {'error_message': 'You do not have permission to access the admin page.'})
        else:
            failed_attempt, created = FailedLoginAttempts.objects.get_or_create(device_id=device_id)
            if not created:
                failed_attempt.attempts += 1
                failed_attempt.save()
                if failed_attempt.attempts >= MAX_FAILED_ATTEMPTS:
                    failed_attempt.is_active = True
                    failed_attempt.save()
                    return render(request, 'adminlogin.html', {'blocked': True, 
                                                            'error_message': 'Your device is permanently blocked due to multiple failed login attempts.'})
            else:
                failed_attempt.attempts = 1
                failed_attempt.save()
            return render(request, 'adminlogin.html', {'error_message': 'Invalid username or password'})
    else:
        response = render(request, 'adminlogin.html')
        if 'device_id' not in request.COOKIES:
            device_id = get_device_id(request)
            set_device_cookie(response, device_id)
        return response
@login_required(login_url="adminlogin")
def admin_logout_view(request):
    logout(request)
    return redirect('adminlogin')

@login_required(login_url='adminlogin')
def adminpage(request):
    return render(request,'adminpage.html')

@login_required(login_url='adminlogin')
@csrf_exempt
def createvisitor(request):
    if request.method == 'POST':
        person_name = request.POST.get('person_name')
        person_phone = request.POST.get('person_phone')
        selected_purpose = request.POST.get('purpose')  
        if selected_purpose == 'Others': 
            purpose = request.POST.get('other-reason')  
        else:
            purpose = selected_purpose 
        print(purpose)
        try:

            visitor = Register(
                person_name=person_name,
                phone_no=person_phone,
                purpose_of_visit=purpose,
                visite_time=timezone.now() 
            )
            visitor.save()
            visit_time_local = visitor.visite_time.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S')
            return JsonResponse({
                'success': True,
                'visitor_data': {
                    'name': visitor.person_name,
                    'phone': str(visitor.phone_no),
                    'visit_id': visitor.visit_id,
                    'visit_time': visit_time_local
                },
                'message': 'Visitor registered successfully!'
            })

        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

    return redirect('adminpage')

@login_required(login_url='adminlogin')
@csrf_exempt
def adminscan(request):
    return render(request,'adminscan.html')

@login_required(login_url='adminlogin')
@csrf_exempt
def update_time(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            qr_data = data.get('qr_data') 

            if qr_data:
                data_parts = [line.split(": ") for line in qr_data.strip().split("\n")]
                visit_id = data_parts[-1][-1].strip() if data_parts else None
                if visit_id:
                    try:
                        print(visit_id)
                        visitor = Register.objects.get(visit_id=visit_id)
                        if visitor.return_time is None:
                            visitor.return_time = timezone.localtime()
                            visitor.save()
                            return JsonResponse({'success': True, 'message': 'Return time updated successfully!'})
                        else:
                            return JsonResponse({'success': False, 'message': 'Return time already updated for this visitor.'})
                    except Register.DoesNotExist:
                        return JsonResponse({'success': False, 'message': 'Visitor not found.'})
                else:
                    return JsonResponse({'success': False, 'message': 'Visit ID not found in QR data.'})
            else:
                return JsonResponse({'success': False, 'message': 'QR data is empty.'})

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data.'})

    return JsonResponse({'success': False, 'message': 'Invalid request method.'})

@login_required(login_url='adminlogin')
@csrf_exempt
def adminvisitors(request):
    return render(request,'adminvisitors.html')

@login_required(login_url='adminlogin')
@csrf_exempt
def get_visitors(request):
    if request.method == 'POST':
        registers = Register.objects.all()

        register_data = []
        for register in registers:
            visit_time_local = register.visite_time.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S')
            if register.return_time:
                return_local_time=register.return_time.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S')
            else:
                return_local_time="Not updated"
            if register.phone_no:
                phone_number=format_number(
                    register.phone_no, 
                    PhoneNumberFormat.NATIONAL
                ).lstrip('0').strip() 
            register_data.append({
                'visit_id': register.visit_id,
                'person_name': register.person_name,
                'phone_no': phone_number,
                'purpose_of_visit': register.purpose_of_visit,
                'visite_time': visit_time_local,
                'return_time': return_local_time
            })

        return JsonResponse({'success': True, 'register_data': register_data})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def handler404(request, exception):
    print("404 handler reached2")
    return render(request, '404.html', status=404)

@login_required(login_url='adminlogin')  
@csrf_exempt
def download_overall_report(request):
    if request.method == 'POST':
        try:
            registers = Register.objects.all()
            
            register_data = []
            for register in registers:
                visit_time_local = register.visite_time.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S')
                if register.return_time:
                    return_local_time=register.return_time.astimezone(timezone.get_current_timezone()).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    return_local_time="Not updated"
                if register.phone_no:
                    phone_number=format_number(
                        register.phone_no, 
                        PhoneNumberFormat.NATIONAL
                    ).lstrip('0').strip() 
                register_data.append({
                    'visit_id': register.visit_id,
                    'person_name': register.person_name,
                    'phone_no': phone_number,
                    'purpose_of_visit': register.purpose_of_visit,
                    'visite_time': visit_time_local,
                    'return_time': return_local_time
                })
            wb = Workbook()
            ws = wb.active
            ws.title = "Registers Data" 
            headers = [
                'Visit ID',
                'Person Name',
                'Phone Number',
                'Purpose of Visit',
                'Visit Time',
                'Return Time',
            ]
            for col_num, header in enumerate(headers, 1):  
                ws.cell(row=1, column=col_num, value=header)
                ws.column_dimensions[chr(ord('A') + col_num - 1)].width = 20 
            for row_num, register in enumerate(register_data, 2):  
                ws.cell(row=row_num, column=1, value=register['visit_id'])
                ws.cell(row=row_num, column=2, value=register['person_name'])
                ws.cell(row=row_num, column=3, value=register['phone_no'])
                ws.cell(row=row_num, column=4, value=register['purpose_of_visit'])
                ws.cell(row=row_num, column=5, value=register['visite_time'])
                ws.cell(row=row_num, column=6, value=register['return_time'])                
            output = BytesIO()
            wb.save(output)  
            output.seek(0) 

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = 'attachment; filename=All_Registers_Report.xlsx'

            return response
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data.'}, status=400)

        except Exception as e: 
            return JsonResponse({'error': f'An error occurred: {str(e)}'}, status=500)
    else:
        return JsonResponse({'error': 'Method not allowed. Use POST.'}, status=405)